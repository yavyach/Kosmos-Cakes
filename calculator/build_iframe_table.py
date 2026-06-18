"""
Берёт тортоначинки.xlsx, добавляет 2 колонки:
  - URL  (прямая ссылка на калькулятор торта)
  - INLINE HTML (готовый сниппет для вставки в редимаг / передачи верстальщику)

Сохраняет рядом как тортоначинки_with_inline.xlsx.
"""

import sys, io, re, shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).parent
# исходный xlsx лежит в C:\Users\A\Desktop\kosmos\, а текущий __file__ —
# в C:\Users\A\Desktop\kosmos\kosmos_calc\calculator\build_iframe_table.py
# поэтому поднимаемся на 2 уровня (kosmos_calc → kosmos)
SRC  = ROOT.parent.parent / "tortonachinki.xlsx"
DST  = ROOT.parent.parent / "тортоначинки_with_inline.xlsx"

GITHUB_CALC = "https://yavyach.github.io/Kosmos-Cakes/calculator"

# Маппинг: имя в xlsx → (тип папки, id файла) ─────────────────────
CAKE_MAP = {
    "богемская рапсодия":      ("weight", "bohemian-rhapsody"),
    "fairy cake":              ("tiered", "fairy-cake"),
    "big cherry fairy cake":   ("tiered", "big-cherry-fairy-cake"),
    "вавилон":                 ("weight", "babylon"),
    "лебединое озеро":         ("weight", "swan-lake"),
    "фаберже":                 ("fixed",  "faberge"),
    "белль":                   ("tiered", "bell"),
    "люмьер":                  ("tiered", "lumiere"),
    "green day":               ("weight", "green-day"),
    "вишневый сад":            ("weight", "cherry-orchard"),
    "куинджи":                 ("tiered", "kuinji"),
    "дарси":                   ("tiered", "darcy"),
    "shine bright":            ("tiered", "shine-bright"),
    "любовное настроение":     ("weight", "love-in-mood"),
    "тирамису":                ("weight", "tiramisu"),
    "secret garden":           ("fixed",  "secret-garden"),
    "анна":                    ("weight", "anna"),
    "шапито":                  ("tiered", "chapito"),
    "келли":                   ("tiered", "kelly"),
    "ягодные поля навсегда":   ("weight", "berry-fields"),
    "вам письмо":              ("weight", "letter"),
    "орфей":                   ("weight", "orpheus"),
    "blueberry hill":          ("fixed",  "blueberry-hill"),
    "аполлон":                 ("weight", "apollo"),
    "сэйлор мун":              ("weight", "sailor-moon"),
    "sailor moon":             ("weight", "sailor-moon"),
    "dancing queen":           ("weight", "dancing-queen"),
    "тоторо":                  ("weight", "totoro"),
    "элизабет":                ("tiered", "elizabeth"),
    # дополнительно из docx (нет в xlsx, но генерируются)
    "la la land":              ("fixed",  "la-la-land"),
    "fuji":                    ("weight", "fuji"),
}

def normalize(name):
    if not name: return ""
    return re.sub(r"\s+", " ", str(name).strip().lower())

def lookup(name):
    n = normalize(name)
    for key, val in CAKE_MAP.items():
        if normalize(key) == n:
            return val
    return None

def _load_cakes():
    bc = ROOT / "build_cakes.py"
    src = bc.read_text(encoding="utf-8")
    cut = src.index("#  ГЕНЕРАЦИЯ")
    ns: dict = {}
    exec(compile(src[:cut], str(bc), "exec"), ns)
    return ns["CAKES"], ns


# ───────────────────────────────────────────────────────────────
shutil.copyfile(SRC, DST)
wb = load_workbook(DST)
ws = wb.active

# Шапка таблицы — первая строка (там колонки начинок). Добавляем 4 справа.
last_col = ws.max_column

URL_COL      = last_col + 1
INLINE_COL   = last_col + 2
URL_M_COL    = last_col + 3
INLINE_M_COL = last_col + 4

ws.cell(row=1, column=URL_COL,      value="URL десктоп")
ws.cell(row=1, column=INLINE_COL,   value="INLINE HTML десктоп")
ws.cell(row=1, column=URL_M_COL,    value="URL мобильный")
ws.cell(row=1, column=INLINE_M_COL, value="INLINE HTML мобильный")

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="D2363C")
for col in (URL_COL, INLINE_COL, URL_M_COL, INLINE_M_COL):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

cakes_list, bc_ns = _load_cakes()
cakes_by_id = {c["id"]: c for c in cakes_list}
css_v = str(int((ROOT / "style.css").stat().st_mtime)) if (ROOT / "style.css").exists() else ""
js_v = str(int((ROOT / "core.js").stat().st_mtime)) if (ROOT / "core.js").exists() else ""

found, missing = [], []
for row_idx in range(2, ws.max_row + 1):
    name = ws.cell(row=row_idx, column=1).value
    if not name or not str(name).strip():
        continue
    match = lookup(name)
    if match:
        type_, cid = match
        url   = f"{GITHUB_CALC}/cakes/{type_}/{cid}.html"
        url_m = f"{GITHUB_CALC}/cakes/{type_}/{cid}.html?ctx=mobile"
        cake = cakes_by_id.get(cid)
        snip = bc_ns["calc_embed_snippet"](cake, css_v, js_v) if cake else ""
        ws.cell(row=row_idx, column=URL_COL,      value=url)
        ws.cell(row=row_idx, column=INLINE_COL,   value=snip)
        ws.cell(row=row_idx, column=URL_M_COL,    value=url_m)
        ws.cell(row=row_idx, column=INLINE_M_COL, value=snip)
        for col in (URL_COL, INLINE_COL, URL_M_COL, INLINE_M_COL):
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="center")
        found.append(str(name).strip())
    else:
        ws.cell(row=row_idx, column=INLINE_COL,   value="— (нет данных по торту)")
        ws.cell(row=row_idx, column=INLINE_M_COL, value="— (нет данных по торту)")
        missing.append(str(name).strip())

# Ширина новых колонок
from openpyxl.utils import get_column_letter
ws.column_dimensions[get_column_letter(URL_COL)].width      = 55
ws.column_dimensions[get_column_letter(INLINE_COL)].width   = 80
ws.column_dimensions[get_column_letter(URL_M_COL)].width    = 55
ws.column_dimensions[get_column_letter(INLINE_M_COL)].width = 80

wb.save(DST)
print(f"Сохранено: {DST}")
print(f"\nСопоставлено: {len(found)} тортов")
for n in found:
    print(f"  ✓ {n}")
if missing:
    print(f"\nНе найдено в калькуляторе: {len(missing)}")
    for n in missing:
        print(f"  ✗ {n}")
